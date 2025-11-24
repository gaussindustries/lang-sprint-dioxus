import json
from pathlib import Path
from openpyxl import load_workbook

INPUT_XLSX = Path("746 most common words.xlsx")  # update if different
OUTPUT_JSON = Path("langs/georgian/test.json")

def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_XLSX}")

    wb = load_workbook(INPUT_XLSX, data_only=True)

    print("Sheets in workbook:", wb.sheetnames)

    entries = []
    rank = 1

    # Go through *all* sheets, just in case
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Scanning sheet: {sheet_name}")

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            # Grab all non-empty cells as strings
            cells = [
                str(cell).strip()
                for cell in row
                if cell is not None and str(cell).strip()
            ]

            # We expect at least 2 non-empty cells: English + Georgian
            if len(cells) < 2:
                continue

            # If you know English is ALWAYS first and Georgian second:
            en_str = cells[0]
            ge_str = cells[1]

            # Quick heuristic: ignore rows that look like headers
            # Adjust these checks as needed.
            if row_idx == 1 and (
                en_str.lower() in ("english", "en") or ge_str.lower() in ("georgian", "ka", "ge")
            ):
                print(f"Skipping header row on sheet {sheet_name}: {cells}")
                continue

            entries.append({
                "rank": rank,
                "ge": ge_str,
                "en": en_str,
            })
            rank += 1

    print(f"Collected {len(entries)} entries before writing.")

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(entries)} entries to {OUTPUT_JSON}")

if __name__ == "__main__":
    main()
